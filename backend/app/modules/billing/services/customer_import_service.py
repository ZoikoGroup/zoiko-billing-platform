"""
modules/billing/services/customer_import_service.py
---------------------------------------------------
Bulk Customer Import Service — mirrors the product bulk import flow.

Supports:
  - CSV and XLSX file parsing
  - Column auto-detection and user-defined mapping
  - Validation of all customer fields (org-scoped, no cross-tenant leakage)
  - Duplicate detection scoped to current organization (customer_code / email)
  - Preview with short-lived TTL cache (30 min, powered by cachetools)
  - Transactional confirm with partial-success result reporting
  - Template generation (CSV + XLSX) with accepted-values annotations
  - Full audit logging for every import operation
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from cachetools import TTLCache
from sqlalchemy.orm import Session

from app.modules.billing.models import BillingAuditAction, BillingCustomer
from app.modules.billing.repositories.customer import CustomerRepository
from app.modules.billing.services.audit_service import BillingAuditService
from app.modules.billing.services.base import filter_allowed
from app.modules.billing.services.customer_service import (
    CUSTOMER_ALLOWED_FIELDS,
    CustomerService,
)
from app.modules.billing.utils.currency_utils import VALID_CURRENCY_CODES as VALID_CURRENCIES

logger = logging.getLogger("zoiko_billing")

# ---------------------------------------------------------------------------
# Supported file columns → customer field mapping
# ---------------------------------------------------------------------------

FIELD_ALIASES: Dict[str, str] = {
    # identity / required
    "customer code": "customer_code",
    "customer_code": "customer_code",
    "code": "customer_code",
    "customer id": "customer_code",
    "reference": "customer_code",
    "company name": "company_name",
    "company": "company_name",
    "customer name": "company_name",
    "organization": "company_name",
    "account name": "company_name",
    # display / names
    "display name": "display_name",
    "display_name": "display_name",
    "legal name": "legal_name",
    "legal_name": "legal_name",
    "first name": "first_name",
    "first_name": "first_name",
    "last name": "last_name",
    "last_name": "last_name",
    # contact
    "email": "email",
    "email address": "email",
    "alternate email": "alternate_email",
    "alternate_email": "alternate_email",
    "secondary email": "alternate_email",
    "mobile": "mobile",
    "mobile number": "mobile",
    "mobile phone": "mobile",
    "cell": "mobile",
    "phone": "phone",
    "telephone": "phone",
    "phone number": "phone",
    "telephone number": "phone",
    "website": "website",
    "web": "website",
    "web site": "website",
    # business profile
    "designation": "designation",
    "job title": "designation",
    "industry": "industry",
    "employee count": "employee_count",
    "employees": "employee_count",
    "no of employees": "employee_count",
    "type": "customer_type",
    "customer type": "customer_type",
    "account type": "customer_type",
    "status": "status",
    "account status": "status",
    # address
    "billing address": "billing_address",
    "address": "billing_address",
    "street address": "billing_address",
    "shipping address": "shipping_address",
    "shipping_address": "shipping_address",
    "delivery address": "shipping_address",
    "billing country": "billing_country",
    "country": "billing_country",
    "billing_country": "billing_country",
    "shipping country": "shipping_country",
    "shipping_country": "shipping_country",
    # billing terms
    "payment terms": "payment_terms",
    "payment_terms": "payment_terms",
    "terms": "payment_terms",
    "currency": "currency",
    "currency code": "currency",
    "credit limit": "credit_limit",
    "credit_limit": "credit_limit",
    "credit days": "credit_days",
    "credit_days": "credit_days",
    "price list": "price_list",
    "price_list": "price_list",
    # tax
    "gst number": "gst_number",
    "gst": "gst_number",
    "gstin": "gst_number",
    "vat number": "vat_number",
    "vat": "vat_number",
    "pan": "pan",
    "pan number": "pan",
    "tin": "tin",
    "tin number": "tin",
    "tax category": "tax_category",
    "tax_category": "tax_category",
    "tax id": "tax_id",
    "tax_id": "tax_id",
    "tax id type": "tax_id_type",
    "tax_id_type": "tax_id_type",
    # misc
    "notes": "notes",
    "remarks": "notes",
    "tags": "tags",
}

# Enterprise-scale safety limits — same rationale as the product importer.
MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 20_000

VALID_CUSTOMER_TYPES = {"business", "individual", "non_profit", "government"}
VALID_STATUSES = {"active", "inactive", "suspended", "closed"}
VALID_PAYMENT_TERMS = {
    "due_on_receipt", "net_15", "net_30", "net_45", "net_60", "net_90",
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# ---------------------------------------------------------------------------
# Global TTL Cache — keyed by (session_id, organization_id)
# Expires after 30 minutes automatically
# ---------------------------------------------------------------------------

_PREVIEW_CACHE: TTLCache = TTLCache(maxsize=512, ttl=1800)  # 30-minute TTL


# ---------------------------------------------------------------------------
# Result data classes (plain dicts; Pydantic schemas are in schemas.py)
# ---------------------------------------------------------------------------

def _row_result(
    row_index: int,
    raw: Dict[str, Any],
    status: str,  # "valid" | "duplicate" | "invalid" | "warning"
    errors: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
    mapped: Optional[Dict[str, Any]] = None,
    matched_id: Optional[int] = None,
    matched_code: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "row_index": row_index,
        "raw_data": raw,
        "mapped_data": mapped or {},
        "status": status,
        "errors": errors or [],
        "warnings": warnings or [],
        "matched_existing_id": matched_id,
        "matched_existing_code": matched_code,
    }


# ---------------------------------------------------------------------------
# File parsing helpers
# ---------------------------------------------------------------------------

def _parse_csv(file_bytes: bytes, max_rows: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse CSV bytes → (headers, rows). Stops reading (rather than truncating
    silently) once max_rows is exceeded, so a huge file is never fully
    materialized in memory."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows: List[Dict[str, str]] = []
    for row in reader:
        if max_rows is not None and len(rows) >= max_rows:
            raise ValueError(
                f"This file has more than {max_rows:,} data rows. "
                f"Please split it into smaller files and import them separately."
            )
        rows.append(dict(row))
    return list(headers), rows


def _parse_xlsx(file_bytes: bytes, max_rows: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse XLSX bytes → (headers, rows). Uses openpyxl's read_only (streaming)
    mode and stops iterating once max_rows is exceeded — protects against a
    small-on-disk XLSX that expands to an enormous number of rows/cells."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX import") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            return [], []
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
        rows: List[Dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue  # skip blank rows
            if max_rows is not None and len(rows) >= max_rows:
                raise ValueError(
                    f"This file has more than {max_rows:,} data rows. "
                    f"Please split it into smaller files and import them separately."
                )
            row_dict = {}
            for h, v in zip(headers, raw_row):
                row_dict[h] = str(v).strip() if v is not None else ""
            rows.append(row_dict)
        return headers, rows
    finally:
        wb.close()


def _check_file_size(file_bytes: bytes) -> None:
    size = len(file_bytes)
    if size > MAX_IMPORT_FILE_SIZE_BYTES:
        raise ValueError(
            f"File is too large ({size / (1024 * 1024):.1f} MB). "
            f"The maximum allowed size is {MAX_IMPORT_FILE_SIZE_BYTES // (1024 * 1024)} MB — "
            f"please split it into smaller files and import them separately."
        )


def _auto_map_columns(headers: List[str]) -> Dict[str, str]:
    """Auto-detect file column → customer field mappings."""
    mapping: Dict[str, str] = {}
    for h in headers:
        if not h:
            continue
        key = h.lower().strip()
        canonical = FIELD_ALIASES.get(key) or FIELD_ALIASES.get(key.rstrip(" *").strip())
        if canonical:
            mapping[h] = canonical
    return mapping


# ---------------------------------------------------------------------------
# Value normalisation helpers
# ---------------------------------------------------------------------------

def _normalize_status(val: str) -> Optional[Tuple[str, bool]]:
    """Convert status string → (status, is_active). Returns None if unparseable."""
    v = val.lower().strip()
    if v in {"true", "yes", "1", "enabled"}:
        return "active", True
    if v in {"false", "no", "0", "disabled"}:
        return "inactive", False
    if v in VALID_STATUSES:
        return v, (v == "active")
    return None


def _normalize_decimal(val: str) -> Optional[float]:
    try:
        return float(val.replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def _normalize_int(val: str) -> Optional[int]:
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, AttributeError):
        return None


def _parse_json_field(raw: str) -> Optional[Any]:
    if not raw:
        return None
    try:
        return json.loads(raw)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------

class CustomerImportService:
    """
    Handles the full lifecycle of a bulk customer import:
      parse → preview → (cache) → confirm → audit
    """

    def __init__(self, db: Session):
        self.db = db
        self.repo = CustomerRepository(db)
        self.audit = BillingAuditService(db)

    # ------------------------------------------------------------------
    # STEP 1: Parse file and return raw columns + sample rows
    # ------------------------------------------------------------------

    def parse_file(
        self,
        file_bytes: bytes,
        filename: str,
    ) -> Dict[str, Any]:
        """
        Parse a CSV or XLSX file.
        Returns:
          detected_columns: list of file column headers
          suggested_mapping: auto-detected column → field mapping
          sample_rows: first 5 rows for display
          total_data_rows: row count (excluding header)
        """
        _check_file_size(file_bytes)
        fname_lower = filename.lower()
        if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
            headers, rows = _parse_xlsx(file_bytes, max_rows=MAX_IMPORT_ROWS)
        elif fname_lower.endswith(".csv"):
            headers, rows = _parse_csv(file_bytes, max_rows=MAX_IMPORT_ROWS)
        else:
            raise ValueError(f"Unsupported file format. Please upload a .csv or .xlsx file.")

        suggested_mapping = _auto_map_columns(headers)

        return {
            "detected_columns": headers,
            "suggested_mapping": suggested_mapping,
            "sample_rows": rows[:5],
            "total_data_rows": len(rows),
        }

    # ------------------------------------------------------------------
    # STEP 2: Validate + preview — returns session token + preview result
    # ------------------------------------------------------------------

    def preview_import(
        self,
        file_bytes: bytes,
        filename: str,
        column_map: Dict[str, str],
        organization_id: int,
        duplicate_strategy: str = "skip",   # skip | overwrite | create_copy | review
    ) -> Dict[str, Any]:
        """
        Validate all rows, detect duplicates, build preview summary.
        Caches the parsed+validated rows under a session_id (30 min TTL).
        """
        _check_file_size(file_bytes)
        fname_lower = filename.lower()
        if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
            headers, raw_rows = _parse_xlsx(file_bytes, max_rows=MAX_IMPORT_ROWS)
        elif fname_lower.endswith(".csv"):
            headers, raw_rows = _parse_csv(file_bytes, max_rows=MAX_IMPORT_ROWS)
        else:
            raise ValueError("Unsupported file format")

        # Auto-detect common/template columns, then let explicit user overrides win.
        effective_map = dict(_auto_map_columns(headers))
        effective_map.update(column_map or {})

        customer_svc = CustomerService(self.db)
        org_currency = customer_svc._resolve_org_currency(organization_id)

        # Pre-fetch existing customers once — used for org-scoped duplicate detection.
        existing_by_code, existing_by_email = self._get_existing_maps(organization_id)

        mapped_rows = []
        for i, raw_row in enumerate(raw_rows):
            row_num = i + 1
            mapped_rows.append((
                row_num,
                raw_row,
                *self._map_and_validate_row(
                    raw=raw_row,
                    column_map=effective_map,
                    row_index=row_num,
                    organization_id=organization_id,
                    org_currency=org_currency,
                ),
            ))

        processed: List[Dict[str, Any]] = []
        counts = {"valid": 0, "duplicate": 0, "invalid": 0, "warning": 0}

        # Within-file duplicate tracking (customer_code / email)
        seen_codes: Dict[str, int] = {}
        seen_emails: Dict[str, int] = {}

        for row_num, raw_row, mapped, errors, warnings in mapped_rows:

            if errors:
                processed.append(_row_result(row_num, raw_row, "invalid", errors=errors, warnings=warnings, mapped=mapped))
                counts["invalid"] += 1
                continue

            code = mapped.get("customer_code")
            email = (mapped.get("email") or "").lower()
            dup_note = None
            matched_id = None
            matched_code = None

            # 1) Existing customer in this org
            existing = existing_by_code.get(code) if code else None
            if existing:
                dup_note = f"Duplicate customer code '{code}'"
                matched_id = existing.id
                matched_code = existing.customer_code
            if not existing and email:
                existing = existing_by_email.get(email)
                if existing:
                    dup_note = f"Duplicate email '{email}'"
                    matched_id = existing.id
                    matched_code = existing.customer_code

            # 2) Duplicate within the same file
            if not existing and code and code in seen_codes:
                dup_note = f"Duplicate customer code '{code}' appears earlier in this file (row {seen_codes[code]})"
            if not existing and not dup_note and email and email in seen_emails:
                dup_note = f"Duplicate email '{email}' appears earlier in this file (row {seen_emails[email]})"

            if not existing:
                if code:
                    seen_codes[code] = row_num
                if email:
                    seen_emails[email] = row_num

            if dup_note:
                processed.append(_row_result(
                    row_num, raw_row, "duplicate",
                    warnings=[dup_note] + warnings,
                    mapped=mapped,
                    matched_id=matched_id,
                    matched_code=matched_code,
                ))
                counts["duplicate"] += 1
            elif warnings:
                processed.append(_row_result(row_num, raw_row, "warning", warnings=warnings, mapped=mapped))
                counts["warning"] += 1
            else:
                processed.append(_row_result(row_num, raw_row, "valid", mapped=mapped))
                counts["valid"] += 1

        session_id = str(uuid.uuid4())
        expires_at = datetime.utcnow().isoformat() + "Z+1800s"

        _PREVIEW_CACHE[(session_id, organization_id)] = {
            "rows": processed,
            "raw_rows": raw_rows,
            "column_map": effective_map,
            "duplicate_strategy": duplicate_strategy,
            "organization_id": organization_id,
        }

        return {
            "session_id": session_id,
            "expires_at": expires_at,
            "total": len(raw_rows),
            "valid": counts["valid"],
            "duplicate": counts["duplicate"],
            "invalid": counts["invalid"],
            "warning": counts["warning"],
            "rows": processed,
            "summary_stats": counts,
        }

    # ------------------------------------------------------------------
    # STEP 3: Confirm import (transactional, partial-success model)
    # ------------------------------------------------------------------

    def confirm_import(
        self,
        session_id: str,
        organization_id: int,
        user_id: int,
        duplicate_strategy: str = "skip",       # global: skip | overwrite | create_copy
        per_row_actions: Optional[Dict[int, str]] = None,  # {row_index: action} for review mode
        offset: int = 0,
        batch_size: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Commit the import.

        Partial-success model:
          - Each row is processed independently in its own transaction.
          - Failures accumulate in 'failed' list but do not abort the batch.
          - The caller sees a detailed per-row result.
          - Audit events are logged for the batch start + end.

        Batching (offset/batch_size): lets a caller process the cached row set
        in slices across multiple requests; each call only processes
        rows[offset:offset+batch_size], the session cache is only evicted once
        the final slice completes, and the response reports whether more
        batches remain so the caller can show real progress.
        """
        cache_key = (session_id, organization_id)
        cached = _PREVIEW_CACHE.get(cache_key)
        if not cached:
            raise ValueError(
                "Import session has expired or is invalid. "
                "Please re-upload your file and preview again."
            )

        if cached["organization_id"] != organization_id:
            raise PermissionError("Session does not belong to this organization.")

        all_rows: List[Dict[str, Any]] = cached["rows"]
        total_rows = len(all_rows)
        rows = all_rows[offset:offset + batch_size] if batch_size is not None else all_rows[offset:]
        batch_end = offset + len(rows)
        is_complete = batch_end >= total_rows

        per_row_actions = per_row_actions or {}
        customer_svc = CustomerService(self.db)

        # Audit: import started (only once, on the first batch)
        if offset == 0:
            self.audit.log(
                organization_id, user_id,
                BillingAuditAction.CREATE, "CustomerImport", None,
                new_values={"session_id": session_id, "total_rows": total_rows, "strategy": duplicate_strategy},
            )

        imported: List[int] = []
        skipped: List[int] = []
        failed: List[Dict[str, Any]] = []
        warning_rows: List[int] = []

        for row in rows:
            row_idx = row["row_index"]
            status = row["status"]
            mapped = row.get("mapped_data", {})
            matched_id = row.get("matched_existing_id")

            if status == "invalid":
                failed.append({"row": row_idx, "error": "; ".join(row.get("errors", ["Validation failed"]))})
                continue

            if status == "warning":
                warning_rows.append(row_idx)

            if status == "duplicate":
                # Resolve action: per-row override > global strategy
                action = per_row_actions.get(row_idx, duplicate_strategy)
                if action == "skip":
                    skipped.append(row_idx)
                    continue
                elif action == "overwrite" and matched_id:
                    try:
                        self._update_customer_from_mapped(
                            customer_svc, matched_id, organization_id, user_id, mapped,
                        )
                        imported.append(row_idx)
                    except Exception as exc:
                        failed.append({"row": row_idx, "error": str(exc)})
                    continue
                elif action == "create_copy":
                    mapped = self._make_unique_code(mapped, organization_id)
                    # falls through to create below
                else:
                    skipped.append(row_idx)
                    continue

            # Create new customer
            try:
                self._create_customer_from_mapped(
                    customer_svc, organization_id, user_id, mapped,
                )
                imported.append(row_idx)
            except Exception as exc:
                failed.append({"row": row_idx, "error": str(exc)})

        # Only evict the cache once the final batch has been processed
        if is_complete:
            _PREVIEW_CACHE.pop(cache_key, None)

        self.audit.log(
            organization_id, user_id,
            BillingAuditAction.UPDATE, "CustomerImport", None,
            new_values={
                "session_id": session_id,
                "batch_offset": offset,
                "batch_rows": len(rows),
                "is_complete": is_complete,
                "imported": len(imported),
                "skipped": len(skipped),
                "failed": len(failed),
                "warnings": len(warning_rows),
            },
        )

        return {
            "imported": len(imported),
            "skipped": len(skipped),
            "failed": len(failed),
            "warnings": len(warning_rows),
            "imported_row_indices": imported,
            "skipped_row_indices": skipped,
            "failed_details": failed,
            "warning_row_indices": warning_rows,
            "total_rows": total_rows,
            "next_offset": None if is_complete else batch_end,
            "is_complete": is_complete,
        }

    # ------------------------------------------------------------------
    # TEMPLATE GENERATION
    # ------------------------------------------------------------------

    def generate_template(self, fmt: str) -> Tuple[bytes, str]:
        """
        Generate a downloadable import template.
        Returns (file_bytes, mimetype).
        """
        headers = [
            "Customer Code *", "Company Name *", "Display Name", "Legal Name",
            "Email", "Alternate Email", "First Name", "Last Name",
            "Mobile", "Phone", "Website", "Designation", "Industry",
            "Employee Count", "Customer Type", "Status",
            "Billing Address", "Shipping Address", "Billing Country", "Shipping Country",
            "Payment Terms", "Currency", "Credit Limit", "Credit Days", "Price List",
            "GST Number", "VAT Number", "PAN", "TIN", "Tax Category", "Tax ID", "Tax ID Type",
            "Notes", "Tags",
        ]
        example_rows = [
            [
                "CUST-1001", "Acme Corp", "Acme Corp", "Acme Corporation",
                "billing@acme.com", "finance@acme.com", "", "",
                "+1-555-0100", "+1-555-0199", "https://acme.example",
                "Procurement", "Technology", "250",
                "business", "active",
                "100 Market St, San Francisco", "100 Market St, San Francisco",
                "US", "US",
                "net_30", "USD", "25000.00", "30", "Standard",
                "GSTIN1234", "", "", "", "standard", "", "",
                "Key enterprise account", '{"priority":"high"}',
            ],
            [
                "CUST-1002", "Bright Startups Inc", "Bright Startups", "",
                "", "", "Jane", "Doe",
                "+1-555-0200", "", "https://bright.example", "CEO", "Software",
                "12",
                "individual", "active",
                "10 Innovation Drive", "",
                "CA", "",
                "net_15", "CAD", "5000.00", "15", "SMB",
                "", "", "", "", "", "", "",
                "Monthly SaaS customer", "",
            ],
        ]
        notes = [
            [
                "Required. Unique per org.",
                "Required. Company / account name.",
                "Optional. Defaults to Company Name.",
                "Optional.",
                "Optional. Must be a valid email. Duplicate emails are flagged.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional. e.g. https://example.com",
                "Optional.",
                "Optional.",
                "Optional. Integer >= 0.",
                "Optional. business | individual | non_profit | government. Defaults to business.",
                "Optional. active | inactive | suspended | closed. Defaults to active.",
                "Optional.",
                "Optional.",
                "Optional. 2-letter ISO code e.g. US, IN, CA.",
                "Optional.",
                "Optional. due_on_receipt | net_15 | net_30 | net_45 | net_60 | net_90. Defaults to net_30.",
                "Optional. 3-letter code e.g. USD, EUR, INR. Defaults to org currency.",
                "Optional. Numeric >= 0.",
                "Optional. Integer 0–365.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional.",
                "Optional. Free-text notes.",
                "Optional. JSON object, e.g. {\"priority\":\"high\"}",
            ]
        ]

        if fmt == "xlsx":
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "Customers"

            header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            req_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(h) + 4)

            for row_data in example_rows:
                ws.append(row_data)
                for col_idx in range(1, len(row_data) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = req_fill

            ws_notes = wb.create_sheet("Field Notes")
            ws_notes.append(["Field", "Notes / Accepted Values"])
            ws_notes["A1"].font = Font(bold=True)
            ws_notes["B1"].font = Font(bold=True)
            for h, note in zip(headers, notes[0]):
                ws_notes.append([h, note])
            ws_notes.column_dimensions["A"].width = 25
            ws_notes.column_dimensions["B"].width = 80

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:  # csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(example_rows)
            output.write("\n# Field Notes:\n")
            for h, note in zip(headers, notes[0]):
                output.write(f"# {h}: {note}\n")
            return output.getvalue().encode("utf-8"), "text/csv"

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_existing_maps(self, organization_id: int) -> Tuple[Dict[str, BillingCustomer], Dict[str, BillingCustomer]]:
        """Return ({customer_code.lower(): customer}, {email.lower(): customer})
        for all non-deleted org customers — used for duplicate detection."""
        by_code: Dict[str, BillingCustomer] = {}
        by_email: Dict[str, BillingCustomer] = {}
        try:
            for customer in self.repo.list_all(organization_id):
                if customer and customer.customer_code:
                    by_code[customer.customer_code.lower()] = customer
                if customer and customer.email:
                    by_email[customer.email.lower()] = customer
        except Exception:
            logger.warning("Could not pre-fetch existing customers for org %s", organization_id, exc_info=True)
        return by_code, by_email

    def _map_and_validate_row(
        self,
        raw: Dict[str, str],
        column_map: Dict[str, str],
        row_index: int,
        organization_id: int,
        org_currency: str,
    ) -> Tuple[Dict[str, Any], List[str], List[str]]:
        """
        Map raw CSV/XLSX columns to customer fields, validate all values.
        Returns (mapped_data, errors, warnings).
        """
        mapped: Dict[str, Any] = {}
        errors: List[str] = []
        warnings: List[str] = []

        for file_col, customer_field in column_map.items():
            value = raw.get(file_col, "")
            if value is None:
                value = ""
            value = str(value).strip()
            mapped[customer_field] = value

        # --- Required fields ---
        customer_code = mapped.get("customer_code", "")
        if not customer_code:
            errors.append("Row is missing required field: Customer Code")
        if len(customer_code) > 50:
            errors.append("Customer Code exceeds 50 characters")

        company_name = mapped.get("company_name", "")
        if not company_name:
            errors.append("Row is missing required field: Company Name")
        if len(company_name) > 255:
            errors.append("Company Name exceeds 255 characters")

        # --- Customer type ---
        ctype = (mapped.get("customer_type") or "").lower().strip()
        if ctype:
            if ctype not in VALID_CUSTOMER_TYPES:
                errors.append(f"Invalid customer type '{ctype}'. Accepted: {', '.join(sorted(VALID_CUSTOMER_TYPES))}")
            else:
                mapped["customer_type"] = ctype
        else:
            mapped["customer_type"] = "business"
            warnings.append("Customer type not specified — defaulted to 'business'")

        # --- Status → status + is_active ---
        status_raw = (mapped.get("status") or "").strip()
        if status_raw:
            parsed = _normalize_status(status_raw)
            if parsed is None:
                warnings.append(f"Unrecognized status '{status_raw}' — defaulted to active")
                mapped["status"] = "active"
                mapped["is_active"] = True
            else:
                mapped["status"], mapped["is_active"] = parsed
        else:
            mapped["status"] = "active"
            mapped["is_active"] = True

        # --- Currency ---
        currency = (mapped.get("currency") or "").upper().strip()
        if currency:
            if currency not in VALID_CURRENCIES:
                errors.append(f"Invalid currency '{currency}'. Use 3-letter ISO codes like USD, EUR, INR.")
            else:
                mapped["currency"] = currency
        else:
            mapped["currency"] = org_currency

        # --- Payment terms ---
        terms = (mapped.get("payment_terms") or "").lower().strip()
        if terms:
            if terms not in VALID_PAYMENT_TERMS:
                errors.append(
                    f"Invalid payment terms '{terms}'. Accepted: {', '.join(sorted(VALID_PAYMENT_TERMS))}"
                )
            else:
                mapped["payment_terms"] = terms
        else:
            mapped["payment_terms"] = "net_30"

        # --- Numeric fields ---
        for field_name, display_name, check in [
            ("credit_limit", "Credit Limit", lambda v: v >= 0),
            ("credit_days", "Credit Days", lambda v: 0 <= v <= 365),
            ("employee_count", "Employee Count", lambda v: v >= 0),
        ]:
            raw_val = mapped.get(field_name, "")
            if raw_val == "" or raw_val is None:
                mapped.pop(field_name, None)
                continue
            if field_name == "credit_limit":
                numeric = _normalize_decimal(str(raw_val))
            else:
                numeric = _normalize_int(str(raw_val))
            if numeric is None:
                errors.append(f"Invalid numeric value for '{display_name}': '{raw_val}'")
            elif not check(numeric):
                errors.append(
                    f"'{display_name}' must be {'0–365' if field_name == 'credit_days' else '>= 0'}. Got: {raw_val}"
                )
            else:
                mapped[field_name] = numeric

        # --- Email validation ---
        for email_field in ("email", "alternate_email"):
            val = (mapped.get(email_field) or "").strip()
            if val:
                if not EMAIL_RE.match(val):
                    errors.append(f"Invalid email '{val}' in '{email_field}'")
                else:
                    mapped[email_field] = val.lower()

        # --- Soft-contact / tax fields (already stripped; keep as-is) ---

        # --- tags (JSON object) ---
        tags_raw = mapped.get("tags", "")
        if tags_raw:
            parsed = _parse_json_field(tags_raw)
            if parsed is None:
                warnings.append("Tags column is not valid JSON — ignored.")
                mapped.pop("tags", None)
            else:
                mapped["tags"] = parsed

        # --- Display name fallback ---
        if not mapped.get("display_name"):
            mapped["display_name"] = company_name or f"{mapped.get('first_name', '')} {mapped.get('last_name', '')}".strip()

        # Only keep allowed fields
        cleaned = {k: v for k, v in mapped.items() if k in CUSTOMER_ALLOWED_FIELDS}
        cleaned["customer_code"] = customer_code
        cleaned["company_name"] = company_name

        return cleaned, errors, warnings

    def _create_customer_from_mapped(
        self,
        svc: CustomerService,
        organization_id: int,
        user_id: int,
        mapped: Dict[str, Any],
    ) -> BillingCustomer:
        data = {k: v for k, v in mapped.items() if v != "" and v is not None}
        data = filter_allowed(data, CUSTOMER_ALLOWED_FIELDS)
        return svc.create_customer(organization_id=organization_id, created_by=user_id, **data)

    def _update_customer_from_mapped(
        self,
        svc: CustomerService,
        customer_id: int,
        organization_id: int,
        user_id: int,
        mapped: Dict[str, Any],
    ) -> BillingCustomer:
        data = {k: v for k, v in mapped.items() if v != "" and v is not None}
        data = filter_allowed(data, CUSTOMER_ALLOWED_FIELDS)
        return svc.update_customer(
            customer_id=customer_id,
            organization_id=organization_id,
            updated_by=user_id,
            **data,
        )

    def _make_unique_code(self, mapped: Dict[str, Any], organization_id: int) -> Dict[str, Any]:
        """Generate a unique customer_code for 'create_copy' duplicate strategy."""
        base_code = f"{mapped.get('customer_code', 'COPY')}-COPY"
        code = base_code
        suffix = 1
        while self.repo.exists(organization_id, customer_code=code):
            suffix += 1
            code = f"{base_code}-{suffix}"
        return {**mapped, "customer_code": code}
