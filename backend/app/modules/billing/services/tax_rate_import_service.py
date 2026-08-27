"""
modules/billing/services/tax_rate_import_service.py
-----------------------------------------------------
Bulk Tax Rate Import Service

Supports:
  - CSV and XLSX file parsing
  - Column auto-detection and user-defined mapping
  - Validation of all tax rate fields (rate range, tax type, currency,
    country, effective-date ordering)
  - Duplicate detection scoped to current organization, keyed by `code`
    (matches the DB's uq_tax_rates_org_code unique constraint)
  - Preview with short-lived TTL cache (30 min, powered by cachetools)
  - Transactional confirm with partial-success result reporting, reusing
    TaxService.create_tax_rate/update_tax_rate so the is_default-per-currency
    uniqueness fix and audit logging apply per row exactly as they do for a
    manually created/edited rate
  - Template generation (CSV + XLSX) with accepted-values annotations
  - No cross-tenant leakage: all DB queries scoped to organization_id

Structured to mirror product_import_service.py -- same session/preview/
confirm shape, same generic ImportPreviewResult/ImportConfirmRequest/
ImportSummaryResult schemas, so the frontend wizard and this service behave
identically to the already-working Customer/Product importers.
"""

from __future__ import annotations

import csv
import io
import logging
import uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from cachetools import TTLCache
from sqlalchemy.orm import Session

from app.modules.billing.models import BillingAuditAction, TaxApplicability, TaxType
from app.modules.billing.repositories.tax import TaxRateRepository
from app.modules.billing.services.audit_service import BillingAuditService
from app.modules.billing.services.product_service import _resolve_org_currency
from app.modules.billing.services.tax_service import TaxService
from app.modules.billing.utils.currency_utils import VALID_CURRENCY_CODES

logger = logging.getLogger("zoiko_billing")

# ---------------------------------------------------------------------------
# Supported file columns → tax rate field mapping
# ---------------------------------------------------------------------------

FIELD_ALIASES: Dict[str, str] = {
    # name
    "name": "name",
    "tax name": "name",
    "tax rate name": "name",
    "title": "name",
    # code
    "code": "code",
    "tax code": "code",
    "rate code": "code",
    "reference": "code",
    # rate
    "rate": "rate",
    "rate (%)": "rate",
    "tax rate": "rate",
    "percentage": "rate",
    "rate %": "rate",
    # tax type
    "tax_type": "tax_type",
    "tax type": "tax_type",
    "type": "tax_type",
    # jurisdiction
    "jurisdiction": "jurisdiction",
    "region": "jurisdiction",
    "state": "jurisdiction",
    # country
    "country_code": "country_code",
    "country": "country_code",
    "country code": "country_code",
    # currency
    "currency_code": "currency_code",
    "currency": "currency_code",
    "currency code": "currency_code",
    # applies to
    "applies_to": "applies_to",
    "applies to": "applies_to",
    "applicability": "applies_to",
    # tax type label
    "tax_type_label": "tax_type_label",
    "tax type label": "tax_type_label",
    "label": "tax_type_label",
    # compound / recoverable / active / default
    "is_compound": "is_compound",
    "is compound": "is_compound",
    "compound": "is_compound",
    "is_recoverable": "is_recoverable",
    "is recoverable": "is_recoverable",
    "recoverable": "is_recoverable",
    "is_active": "is_active",
    "status": "is_active",
    "active": "is_active",
    "is active": "is_active",
    "is_default": "is_default",
    "is default": "is_default",
    "default": "is_default",
    # priority
    "priority": "priority",
    # dates
    "effective_from": "effective_from",
    "effective from": "effective_from",
    "start date": "effective_from",
    "effective_to": "effective_to",
    "effective until": "effective_to",
    "effective to": "effective_to",
    "end date": "effective_to",
}

REQUIRED_FIELDS = {"name", "code", "jurisdiction", "rate", "tax_type"}

# Enterprise-scale safety limits -- mirrors product_import_service.py so a
# huge or maliciously crafted file can never be fully materialized in memory.
MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 20_000

VALID_TAX_TYPES = {t.value for t in TaxType}
VALID_APPLIES_TO = {a.value for a in TaxApplicability}

# ---------------------------------------------------------------------------
# Global TTL Cache — keyed by (session_id, organization_id)
# ---------------------------------------------------------------------------

_PREVIEW_CACHE: TTLCache = TTLCache(maxsize=512, ttl=1800)  # 30-minute TTL


def _row_result(
    row_index: int,
    raw: Dict[str, Any],
    status: str,  # "valid" | "duplicate" | "invalid" | "warning"
    errors: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
    mapped: Optional[Dict[str, Any]] = None,
    matched_id: Optional[int] = None,
    matched_code: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "row_index": row_index,
        "raw_data": raw,
        "mapped_data": mapped or {},
        "status": status,
        "errors": errors or [],
        "warnings": warnings or [],
        "matched_existing_id": matched_id,
        "matched_existing_code": matched_code,
    }


# ---------------------------------------------------------------------------
# File parsing helpers (identical shape to product_import_service.py)
# ---------------------------------------------------------------------------

def _parse_csv(file_bytes: bytes, max_rows: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows: List[Dict[str, str]] = []
    for row in reader:
        if max_rows is not None and len(rows) >= max_rows:
            raise ValueError(
                f"This file has more than {max_rows:,} data rows. "
                f"Please split it into smaller files and import them separately."
            )
        rows.append(dict(row))
    return list(headers), rows


def _parse_xlsx(file_bytes: bytes, max_rows: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX import") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            return [], []
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
        rows: List[Dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            if max_rows is not None and len(rows) >= max_rows:
                raise ValueError(
                    f"This file has more than {max_rows:,} data rows. "
                    f"Please split it into smaller files and import them separately."
                )
            row_dict = {}
            for h, v in zip(headers, raw_row):
                row_dict[h] = str(v).strip() if v is not None else ""
            rows.append(row_dict)
        return headers, rows
    finally:
        wb.close()


def _check_file_size(file_bytes: bytes) -> None:
    size = len(file_bytes)
    if size > MAX_IMPORT_FILE_SIZE_BYTES:
        raise ValueError(
            f"File is too large ({size / (1024 * 1024):.1f} MB). "
            f"The maximum allowed size is {MAX_IMPORT_FILE_SIZE_BYTES // (1024 * 1024)} MB — "
            f"please split it into smaller files and import them separately."
        )


def _auto_map_columns(headers: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for h in headers:
        if not h:
            continue
        key = h.lower().strip()
        canonical = FIELD_ALIASES.get(key) or FIELD_ALIASES.get(key.rstrip(" *").strip())
        if canonical:
            mapping[h] = canonical
    return mapping


# ---------------------------------------------------------------------------
# Value normalisation helpers
# ---------------------------------------------------------------------------

def _normalize_bool(val: str) -> Optional[bool]:
    v = val.lower().strip()
    if v in {"true", "yes", "1", "active", "enabled"}:
        return True
    if v in {"false", "no", "0", "inactive", "disabled"}:
        return False
    return None


def _normalize_decimal(val: str) -> Optional[float]:
    try:
        return float(val.replace(",", "").replace("%", "").strip())
    except (ValueError, AttributeError):
        return None


def _normalize_date(val: str) -> Tuple[Optional[date], bool]:
    """Returns (parsed_date_or_None, was_parseable). A blank string parses to
    (None, True) -- "no value supplied" is not a parse failure."""
    v = (val or "").strip()
    if not v:
        return None, True
    try:
        return date.fromisoformat(v[:10]), True
    except ValueError:
        return None, False


class TaxRateImportService:
    """Handles the full lifecycle of a bulk tax rate import:
    parse → preview → (cache) → confirm → audit."""

    def __init__(self, db: Session):
        self.db = db
        self.rate_repo = TaxRateRepository(db)
        self.tax_svc = TaxService(db)
        self.audit = BillingAuditService(db)

    # ------------------------------------------------------------------
    # STEP 1: Parse file and return raw columns + sample rows
    # ------------------------------------------------------------------

    def parse_file(self, file_bytes: bytes, filename: str) -> Dict[str, Any]:
        _check_file_size(file_bytes)
        fname_lower = filename.lower()
        if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
            headers, rows = _parse_xlsx(file_bytes, max_rows=MAX_IMPORT_ROWS)
        elif fname_lower.endswith(".csv"):
            headers, rows = _parse_csv(file_bytes, max_rows=MAX_IMPORT_ROWS)
        else:
            raise ValueError("Unsupported file format. Please upload a .csv or .xlsx file.")

        suggested_mapping = _auto_map_columns(headers)
        return {
            "detected_columns": headers,
            "suggested_mapping": suggested_mapping,
            "sample_rows": rows[:5],
            "total_data_rows": len(rows),
        }

    # ------------------------------------------------------------------
    # STEP 2: Validate + preview
    # ------------------------------------------------------------------

    def preview_import(
        self,
        file_bytes: bytes,
        filename: str,
        column_map: Dict[str, str],
        organization_id: int,
        duplicate_strategy: str = "skip",
    ) -> Dict[str, Any]:
        _check_file_size(file_bytes)
        fname_lower = filename.lower()
        if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
            headers, raw_rows = _parse_xlsx(file_bytes, max_rows=MAX_IMPORT_ROWS)
        elif fname_lower.endswith(".csv"):
            headers, raw_rows = _parse_csv(file_bytes, max_rows=MAX_IMPORT_ROWS)
        else:
            raise ValueError("Unsupported file format")

        effective_map = dict(_auto_map_columns(headers))
        effective_map.update(column_map or {})

        org_currency = _resolve_org_currency(self.db, organization_id)

        mapped_rows = []
        seen_codes_in_file: Dict[str, int] = {}
        for i, raw_row in enumerate(raw_rows):
            row_num = i + 1
            mapped, errors, warnings = self._map_and_validate_row(
                raw=raw_row,
                column_map=effective_map,
                org_currency=org_currency,
            )
            mapped_rows.append((row_num, raw_row, mapped, errors, warnings))

        existing_matches = self.rate_repo.list_matching_codes(
            organization_id=organization_id,
            codes=[m.get("code") for _, _, m, errs, _ in mapped_rows if not errs and m.get("code")],
        )
        existing_by_code = {r.code: r for r in existing_matches}

        processed: List[Dict[str, Any]] = []
        counts = {"valid": 0, "duplicate": 0, "invalid": 0, "warning": 0}

        for row_num, raw_row, mapped, errors, warnings in mapped_rows:
            if errors:
                processed.append(_row_result(row_num, raw_row, "invalid", errors=errors, warnings=warnings, mapped=mapped))
                counts["invalid"] += 1
                continue

            code = mapped.get("code")
            existing = existing_by_code.get(code)

            # Duplicate within the same file: first occurrence stays valid,
            # later ones are flagged as duplicates of the in-file row.
            in_file_dup = code in seen_codes_in_file
            if code:
                seen_codes_in_file.setdefault(code, row_num)

            if existing or in_file_dup:
                warn_msg = f"Duplicate code '{code}'" + (
                    f" (matches existing rate)" if existing else f" (also used by row {seen_codes_in_file.get(code)})"
                )
                processed.append(_row_result(
                    row_num, raw_row, "duplicate",
                    warnings=[warn_msg] + warnings,
                    mapped=mapped,
                    matched_id=existing.id if existing else None,
                    matched_code=existing.code if existing else None,
                ))
                counts["duplicate"] += 1
            elif warnings:
                processed.append(_row_result(row_num, raw_row, "warning", warnings=warnings, mapped=mapped))
                counts["warning"] += 1
            else:
                processed.append(_row_result(row_num, raw_row, "valid", mapped=mapped))
                counts["valid"] += 1

        session_id = str(uuid.uuid4())
        expires_at = datetime.utcnow().isoformat() + "Z+1800s"

        _PREVIEW_CACHE[(session_id, organization_id)] = {
            "rows": processed,
            "duplicate_strategy": duplicate_strategy,
            "organization_id": organization_id,
        }

        return {
            "session_id": session_id,
            "expires_at": expires_at,
            "total": len(raw_rows),
            "valid": counts["valid"],
            "duplicate": counts["duplicate"],
            "invalid": counts["invalid"],
            "warning": counts["warning"],
            "rows": processed,
            "summary_stats": counts,
        }

    # ------------------------------------------------------------------
    # STEP 3: Confirm import (transactional, partial-success model)
    # ------------------------------------------------------------------

    def confirm_import(
        self,
        session_id: str,
        organization_id: int,
        user_id: int,
        duplicate_strategy: str = "skip",
        per_row_actions: Optional[Dict[int, str]] = None,
        offset: int = 0,
        batch_size: Optional[int] = None,
    ) -> Dict[str, Any]:
        cache_key = (session_id, organization_id)
        cached = _PREVIEW_CACHE.get(cache_key)
        if not cached:
            raise ValueError(
                "Import session has expired or is invalid. "
                "Please re-upload your file and preview again."
            )
        if cached["organization_id"] != organization_id:
            raise PermissionError("Session does not belong to this organization.")

        all_rows: List[Dict[str, Any]] = cached["rows"]
        total_rows = len(all_rows)
        rows = all_rows[offset:offset + batch_size] if batch_size is not None else all_rows[offset:]
        batch_end = offset + len(rows)
        is_complete = batch_end >= total_rows

        per_row_actions = per_row_actions or {}

        if offset == 0:
            self.audit.log(
                organization_id, user_id,
                BillingAuditAction.CREATE, "TaxRateImport", None,
                new_values={"session_id": session_id, "total_rows": total_rows, "strategy": duplicate_strategy},
            )

        imported: List[int] = []
        skipped: List[int] = []
        failed: List[Dict[str, Any]] = []
        warning_rows: List[int] = []

        for row in rows:
            row_idx = row["row_index"]
            status = row["status"]
            mapped = row.get("mapped_data", {})
            matched_id = row.get("matched_existing_id")

            if status == "invalid":
                failed.append({"row": row_idx, "error": "; ".join(row.get("errors", ["Validation failed"]))})
                continue

            if status == "warning":
                warning_rows.append(row_idx)

            if status == "duplicate":
                action = per_row_actions.get(row_idx, duplicate_strategy)
                if action == "skip":
                    skipped.append(row_idx)
                    continue
                elif action == "overwrite" and matched_id:
                    try:
                        self.tax_svc.update_tax_rate(
                            rate_id=matched_id, organization_id=organization_id,
                            updated_by=user_id,
                            **{k: v for k, v in mapped.items() if k not in {"code"}},
                        )
                        imported.append(row_idx)
                    except Exception as exc:
                        failed.append({"row": row_idx, "error": str(exc)})
                    continue
                elif action == "create_copy":
                    mapped = self._make_unique_code(mapped, organization_id)
                else:
                    skipped.append(row_idx)
                    continue

            try:
                self.tax_svc.create_tax_rate(
                    organization_id=organization_id, created_by=user_id, **mapped,
                )
                imported.append(row_idx)
            except Exception as exc:
                failed.append({"row": row_idx, "error": str(exc)})

        if is_complete:
            _PREVIEW_CACHE.pop(cache_key, None)

        self.audit.log(
            organization_id, user_id,
            BillingAuditAction.UPDATE, "TaxRateImport", None,
            new_values={
                "session_id": session_id,
                "batch_offset": offset,
                "batch_rows": len(rows),
                "is_complete": is_complete,
                "imported": len(imported),
                "skipped": len(skipped),
                "failed": len(failed),
                "warnings": len(warning_rows),
            },
        )

        return {
            "imported": len(imported),
            "skipped": len(skipped),
            "failed": len(failed),
            "warnings": len(warning_rows),
            "imported_row_indices": imported,
            "skipped_row_indices": skipped,
            "failed_details": failed,
            "warning_row_indices": warning_rows,
            "total_rows": total_rows,
            "next_offset": None if is_complete else batch_end,
            "is_complete": is_complete,
        }

    # ------------------------------------------------------------------
    # TEMPLATE GENERATION
    # ------------------------------------------------------------------

    def generate_template(self, fmt: str) -> Tuple[bytes, str]:
        headers = [
            "Name *", "Code *", "Tax Type *", "Rate (%) *", "Jurisdiction *",
            "Country Code", "Currency", "Applies To", "Effective From",
            "Effective Until", "Is Compound", "Is Recoverable", "Is Default",
            "Priority", "Status",
        ]
        example_rows = [
            [
                "India GST 5%", "IN-GST-5", "gst", "5", "India",
                "IN", "INR", "both", str(date.today()),
                "", "false", "true", "false", "0", "active",
            ],
            [
                "India GST 18%", "IN-GST-18", "gst", "18", "India",
                "IN", "INR", "both", str(date.today()),
                "", "false", "true", "true", "0", "active",
            ],
        ]
        notes = [
            "Required. Tax rate name.",
            "Required. Unique code per org — used to detect duplicates on re-import.",
            f"Required. Values: {', '.join(sorted(VALID_TAX_TYPES))}",
            "Required. Numeric, 0–100.",
            "Required. Free text, e.g. a country, state, or region name.",
            "Optional. 2-letter ISO code (e.g. IN, US, GB) or a recognized country name.",
            "Optional. 3-letter ISO code (e.g. INR, USD, EUR). Defaults to org currency.",
            f"Optional. Values: {', '.join(sorted(VALID_APPLIES_TO))}. Defaults to 'both'.",
            "Optional. YYYY-MM-DD. Defaults to today.",
            "Optional. YYYY-MM-DD. Must be on/after Effective From if set.",
            "Optional. true/false. Defaults to false.",
            "Optional. true/false. Defaults to true.",
            "Optional. true/false. If multiple rows in one file default the same "
            "currency, the last one processed wins — only one default per "
            "org+currency is ever kept.",
            "Optional. Integer, controls compound-tax processing order. Defaults to 0.",
            "Optional. active | inactive. Defaults to active.",
        ]

        if fmt == "xlsx":
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tax Rates"

            header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            req_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(h) + 4)

            for row_data in example_rows:
                ws.append(row_data)
                for col_idx in range(1, len(row_data) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = req_fill

            ws_notes = wb.create_sheet("Field Notes")
            ws_notes.append(["Field", "Notes / Accepted Values"])
            ws_notes["A1"].font = Font(bold=True)
            ws_notes["B1"].font = Font(bold=True)
            for h, note in zip(headers, notes):
                ws_notes.append([h, note])
            ws_notes.column_dimensions["A"].width = 22
            ws_notes.column_dimensions["B"].width = 85

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(example_rows)
            output.write("\n# Field Notes:\n")
            for h, note in zip(headers, notes):
                output.write(f"# {h}: {note}\n")
            return output.getvalue().encode("utf-8"), "text/csv"

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _map_and_validate_row(
        self,
        raw: Dict[str, str],
        column_map: Dict[str, str],
        org_currency: str,
    ) -> Tuple[Dict[str, Any], List[str], List[str]]:
        mapped: Dict[str, Any] = {}
        errors: List[str] = []
        warnings: List[str] = []

        for file_col, field in column_map.items():
            value = raw.get(file_col, "")
            if value is None:
                value = ""
            mapped[field] = str(value).strip()

        # --- Required fields ---
        for field, display in [
            ("name", "Name"), ("code", "Code"), ("jurisdiction", "Jurisdiction"),
        ]:
            if not mapped.get(field):
                errors.append(f"Row is missing required field: {display}")

        # --- Rate ---
        rate_raw = mapped.get("rate", "")
        if not rate_raw:
            errors.append("Row is missing required field: Rate")
        else:
            rate = _normalize_decimal(str(rate_raw))
            if rate is None:
                errors.append(f"Invalid numeric value for 'Rate': '{rate_raw}'")
            elif rate < 0 or rate > 100:
                errors.append(f"Rate must be between 0 and 100. Got: {rate}")
            else:
                mapped["rate"] = rate

        # --- Tax type ---
        tax_type_raw = (mapped.get("tax_type") or "").lower().strip()
        if not tax_type_raw:
            errors.append("Row is missing required field: Tax Type")
        elif tax_type_raw not in VALID_TAX_TYPES:
            errors.append(f"Invalid tax type '{tax_type_raw}'. Accepted: {', '.join(sorted(VALID_TAX_TYPES))}")
        else:
            mapped["tax_type"] = tax_type_raw

        # --- Applies to ---
        applies_raw = (mapped.get("applies_to") or "").lower().strip()
        if applies_raw:
            if applies_raw not in VALID_APPLIES_TO:
                warnings.append(f"Unrecognized 'Applies To' value '{applies_raw}' — defaulted to 'both'")
                mapped["applies_to"] = "both"
            else:
                mapped["applies_to"] = applies_raw
        else:
            mapped["applies_to"] = "both"

        # --- Currency ---
        currency_raw = (mapped.get("currency_code") or "").upper().strip()
        if currency_raw:
            if currency_raw not in VALID_CURRENCY_CODES:
                errors.append(f"Invalid currency '{currency_raw}'. Use 3-letter ISO codes like USD, EUR, INR.")
            else:
                mapped["currency_code"] = currency_raw
        else:
            mapped["currency_code"] = org_currency

        # --- Country ---
        country_raw = (mapped.get("country_code") or "").strip()
        if country_raw:
            from app.modules.auth.country_currency import resolve_country_code
            resolved = resolve_country_code(country_raw)
            if not resolved:
                errors.append(f"Country '{country_raw}' is not recognized. Use a valid name or 2-letter ISO code.")
            else:
                mapped["country_code"] = resolved
        else:
            mapped.pop("country_code", None)

        # --- Dates ---
        eff_from_raw = mapped.get("effective_from", "")
        eff_from, parseable = _normalize_date(eff_from_raw)
        if not parseable:
            errors.append(f"Invalid date for 'Effective From': '{eff_from_raw}'. Use YYYY-MM-DD.")
        else:
            mapped["effective_from"] = eff_from or date.today()

        eff_to_raw = mapped.get("effective_to", "")
        eff_to, parseable = _normalize_date(eff_to_raw)
        if not parseable:
            errors.append(f"Invalid date for 'Effective Until': '{eff_to_raw}'. Use YYYY-MM-DD.")
        elif eff_to is not None:
            if not errors and eff_to < mapped.get("effective_from", date.today()):
                errors.append("'Effective Until' must be on or after 'Effective From'.")
            mapped["effective_to"] = eff_to
        else:
            mapped.pop("effective_to", None)

        # --- Booleans ---
        for field, display, default in [
            ("is_compound", "Is Compound", False),
            ("is_recoverable", "Is Recoverable", True),
            ("is_default", "Is Default", False),
        ]:
            raw_val = (mapped.get(field) or "").strip()
            if raw_val:
                parsed = _normalize_bool(raw_val)
                if parsed is None:
                    warnings.append(f"Unrecognized value for '{display}': '{raw_val}' — defaulted to {default}")
                    mapped[field] = default
                else:
                    mapped[field] = parsed
            else:
                mapped[field] = default

        status_raw = (mapped.get("is_active") or "").strip()
        if status_raw:
            parsed = _normalize_bool(status_raw)
            if parsed is None:
                warnings.append(f"Unrecognized status '{status_raw}' — defaulted to active")
                mapped["is_active"] = True
            else:
                mapped["is_active"] = parsed
        else:
            mapped["is_active"] = True

        # --- Priority ---
        priority_raw = (mapped.get("priority") or "").strip()
        if priority_raw:
            try:
                mapped["priority"] = int(float(priority_raw))
            except ValueError:
                warnings.append(f"Invalid priority '{priority_raw}' — defaulted to 0")
                mapped["priority"] = 0
        else:
            mapped["priority"] = 0

        # --- Tax type label passthrough (free text, no validation) ---
        if not mapped.get("tax_type_label"):
            mapped.pop("tax_type_label", None)

        return mapped, errors, warnings

    def _make_unique_code(self, mapped: Dict[str, Any], organization_id: int) -> Dict[str, Any]:
        base_code = f"{mapped.get('code', 'COPY')}-COPY"
        code = base_code
        suffix = 1
        while self.rate_repo.exists(organization_id, code=code):
            suffix += 1
            code = f"{base_code}-{suffix}"
        return {**mapped, "code": code}
